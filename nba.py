import io
import openpyxl
import requests
from parsel import Selector
import pandas as pd
import datetime
import logging, coloredlogs
from copy import copy
from playwright.sync_api import TimeoutError, Route
from urllib.parse import urljoin
import re
import dateparser
from playwright.sync_api import sync_playwright
from requests.cookies import RequestsCookieJar


logger = logging.getLogger("__name__")
coloredlogs.install("INFO", logger=logger)

_date = input("Enter date:")
playwright_timeout = 120000





########################## utils ############################


def init_playwright(headless=True):
    play = sync_playwright().start()
    browser = play.firefox.launch_persistent_context(headless=headless, user_data_dir="./program-data", viewport=None, ignore_https_errors=True)
    page = browser.pages[0]
    return (play, browser, page)

def is_exist(response, e):
    sel = Selector(response)
    return sel.xpath(e)

def playwright_cookies_to_requests(cookies):
    cookie_jar = RequestsCookieJar()
    for cookie in cookies:
        cookie_jar.set(cookie['name'], cookie['value'], domain=cookie['domain'])
    return cookie_jar


def convert_cells(filename):
    wb = openpyxl.load_workbook(filename)
    for ws in wb.worksheets:
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell_value = str(cell.value).strip()
                try:
                    cell.value = int(cell_value)
                    continue
                except ValueError:
                    pass
                try:
                    cell.value = float(cell_value)
                    continue
                except ValueError:
                    pass
                if cell_value.endswith('%'):
                    try:
                        cell.value = float(cell_value.rstrip('%')) / 100.0
                        continue
                    except ValueError:
                        pass
                cell.value = cell_value
    wb.save(filename)


def update_template(master_workbook, template_workbook, sheet_name, cols):
    """Updates a sheet in the template workbook using values from a master workbook."""

    master_sheet = master_workbook[sheet_name]
    template_sheet = template_workbook[sheet_name]

    for master_col in master_sheet.iter_cols():
        header_name = master_col[0].value
        col_to_replace = cols.get(header_name)

        if col_to_replace is not None:  # Only update columns present in 'cols'
            if 'paydirt' in sheet_name.lower():
                for idx, row in enumerate(template_sheet.iter_rows(min_col=col_to_replace, max_col=col_to_replace), start=1):
                    for cell in row:
                        cell.value = None
                    if idx == 16:
                        break
            # for row in template_sheet.iter_rows(min_col=col_to_replace, max_col=col_to_replace):
            #     for cell in row:
            #         cell.value = None
            col_values = [cell.value for cell in master_col]
            for i, value in enumerate(col_values, start=1):
                template_sheet.cell(row=i, column=col_to_replace).value = value

    
def to_template(_date):
    master_workbook = openpyxl.load_workbook("master_NBA_file.xlsx")
    template_workbook = openpyxl.load_workbook("C:\\Users\\d_lom\\OneDrive\\NBA\\nba_slate_today.xlsx")
    # template_workbook = openpyxl.load_workbook("template.xlsx")
    # Paydirt
    cols = {
        "Team":3,
        "Opp":4,
        "Team Points":5,
        "Opp Points":6,
        "Proj Total":7,
        "Proj Spread":8,
        "Proj Winner":9,
        "Win%":10
    }
    update_template(master_workbook, template_workbook, 'Paydirt', cols) 
    # AM
    cols = {
        "Date":3,
        "Time":4,
        "Acro":5,
        "Opponent":6,
        "H/R":7,
        "Rest":8,
        "Opp Rest":9,
        "Games last 7":10,
        "Games last 7 Opp":11,
        "Expected Margin":12,
        "Win %":13,
        "Team 1 Injury & Rotation Adjustment":14,
        "Team 2 Injury & Rotation Adjustment":15,
        "Injury and Rotation Adjusted Expected Margin":16,
        "Injury and Rotation Adjusted Win %":17,
        "Total Proj":18,
        "Std Dev":19,
        "Projected Points":20,
        "Team Date":21
    }
    update_template(master_workbook, template_workbook, 'AM', cols) 
    # ActPubSPR
    cols = {
        'Date':2, 
        'Time':3, 
        'Game':4, 
        'Team':5, 
        'Pro-Line':6, 
        'Bet %':7, 
        'Money %':8, 
        'O/U Pro-Line':9, 
        'O/U Bet %':10, 
        'O/U Money %':11
    }
    update_template(master_workbook, template_workbook, 'ActPubSPR', cols) 
    # BPSpread
    cols = {'Team': 2, 'Open': 3,'Spread':4, 'Line':5}
    update_template(master_workbook, template_workbook, 'BPSpread', cols)
    # BPTotal
    cols = {'Team': 2, 'Open': 3, 'O/U': 4, 'Total': 5, 'Line':6}
    update_template(master_workbook, template_workbook, 'BPTotal', cols)
    # PT
    cols = {
        "Home":2,
        "Road":3,
        "Line":4,
        "lineopen":5,
        "Sagarin":6,
        "Sagarin Recent":7,
        "Sagarin Pred.":8,
        "Sagarin Golden Mean":9,
        "Sonny Moore":10,
        "Massey Ratings":11,
        "Stat Fox":12,
        "ESPN FPI":13,
        "Dunkel Index":14,
        "Dokter Entropy":15,
        "Versus Sports Simulator":16,
        "Donchess Inference":17,
        "Talisman Red":18,
        "Roundtable":19,
        "lineavg":20
    }
    update_template(master_workbook, template_workbook, 'PT', cols)
    # DE
    cols = {
        "Winner":2,
        "By":3,
        "Loser":4,
        "%":5,
        "ML":6
    }
    update_template(master_workbook, template_workbook, 'DE', cols)
    # Sangarin
    cols = {
        "Favorite":2,
        "Rating":3,
        "Predict":4,
        "Golden":5,
        "Recent":6,
        "Underdog":7,
        "Odds":8,
        "PCT%":9,
        "Total":10
    }
    update_template(master_workbook, template_workbook, 'Sangarin', cols)
    # Massey
    cols = {
        "Date":2,
        "Time":3,
        "Team":4,
        "Pred":5,
        "Pwin":6,
        "Spread":7,
        "Total":8
    }
    update_template(master_workbook, template_workbook, 'Massey', cols)
    # DRatings
    cols = {
        "Time":2,
        "Teams":3,
        "Win":4,
        "Best ML":5,
        "Best Spread":6,
        "Points":7,
        "Spread":8,
        "Total Points":9,
        "Best O/U":10,
        "Bet Value":13
    }
    update_template(master_workbook, template_workbook, 'DRatings', cols)
    # BetQL
    cols = {
        "Team":2,
        "Score":3,
        "Spread":4,
        "Rating":5,
        "Total":6
    }
    update_template(master_workbook, template_workbook, 'BetQL', cols)
    # Sportsline
    cols = {
        "Team":2,
        "Projected Score":3,
        "Spread":4,
        "Total":5
    }
    update_template(master_workbook, template_workbook, 'Sportsline', cols)
    # SPExperts
    cols = {
        "Team":2,
        "Pick":3,
        "Analysis":4,
        "Expert":5
    }
    update_template(master_workbook, template_workbook, 'SPExperts', cols)  
    # TeamStats, HomeStats, RoadStats
    cols = {
        "": 2,
        "TEAM":3,
        "GP":4,
        "W":5,
        "L":6,
        "MIN":7,
        "OFFRTG":8,
        "DEFRTG":9,
        "NETRTG":10,
        "AST%":11,
        "AST/TO":12,
        "AST":13,
        "OREB%":14,
        "DREB%":15,
        "REB%":16,
        "TOV%":17,
        "EFG%":18,
        "TS%":19,
        "PACE":20,
        "PIE":21,
        "POSS":22
    }
    update_template(master_workbook, template_workbook, 'TeamStats', cols)
    update_template(master_workbook, template_workbook, 'HomeStats', cols)
    update_template(master_workbook, template_workbook, 'RoadStats', cols)

    filename = "C:\\Users\\d_lom\\OneDrive\\NBA\\nba_slate_today.xlsx"
    # filename = "template.xlsx"
    template_workbook.save(filename)





##############################################


class ActionNetwork:
    url = "https://www.actionnetwork.com/nba/projections"

    def login(self):
        try:
            self.page.goto("https://www.actionnetwork.com/login", timeout=playwright_timeout)
            self.page.wait_for_timeout(5000)
            response = self.page.content()
            if not is_exist(response, "//button[contains(@class, 'user-component__button user-component__login')]"):
                raise TimeoutError("Already logged in")
            self.page.locator("//button[contains(@class, 'user-component__button user-component__login')]").click()
            self.page.wait_for_timeout(2000)
            self.page.locator("//input[not(@id) and @name='email']").fill("diego.lomanto+AL@gmail.com")
            self.page.wait_for_timeout(1000)
            self.page.locator("//input[not(@id) and @name='password']").fill("dy04rR9rL5m3")
            self.page.wait_for_timeout(1000)
            self.page.locator("//button[@type='submit' and not(@id)]").click()
            self.page.wait_for_selector("//button[@type='submit' and not(@id)]", state="hidden")
            logger.debug("logged in")
            return True
        except TimeoutError:
            return True
        except Exception as e:
            logger.debug("error in login")
            logger.error(e)
            return False

    @staticmethod
    def parse_date(date_str):
        current_year = datetime.datetime.now().year
        parsed_date = datetime.datetime.strptime(date_str + f" {current_year}", "%a %b %d %Y")
        return parsed_date

    def go_down(self):
        for i in range(90):
            self.page.mouse.wheel(0, i)
        self.page.wait_for_timeout(1000)

    def go_up(self):
        for i in range(90):
            self.page.mouse.wheel(0, -i)
        self.page.wait_for_timeout(1000)

    def find_date(self):
        for _ in range(5): # 5 dates forward or backward
            self.page.wait_for_timeout(5000)
            content = self.page.content()
            page_date = self.parse_date(Selector(text=content).xpath("//span[@class='day-nav__display']/text()").get())
            if page_date != self.user_date:
                logger.debug("date not found")
                # decide to move forward or backward
                if page_date > self.user_date:
                    self.page.locator("//button[@aria-label='Previous Date']").click()
                    logger.debug("going into past")
                else:
                    self.page.locator("//button[@aria-label='Next Date']").click()
                    logger.debug("going into future")
            else:
                logger.debug("date found")
                break

    def scroll_into_view(self, element):
        self.page.locator(element).hover()
        self.go_down()
        self.go_up()

    def parse_spread(self, response):
        items = []
        try:
            sel = Selector(text=response)
            for match in sel.xpath("//table/tbody/tr"):
                date_time = match.xpath("./td[1]/div/div/text()").get('')
                if ':' in date_time:
                    match_date = date_time.split(',')[0] if ',' in date_time else ''
                    match_time = date_time.split(',')[-1] if ',' in date_time else date_time
                else:
                    match_date, match_time = '', ''
                items.append({
                    "Date": match_date,
                    "Time": match_time,
                    "Game": match.xpath(
                        "./td[1]//div[@class='game-info__teams'][1]//div[@class='game-info__team--desktop']/span/text()").get() + ' vs ' + match.xpath(
                        "./td[1]//div[@class='game-info__teams'][2]//div[@class='game-info__team--desktop']/span/text()").get(),
                    "Team": match.xpath(
                        "./td[1]//div[@class='game-info__teams'][1]//div[@class='game-info__team--desktop']/span/text()").get(),
                    "Pro-Line": match.xpath("./td[3]/div/div[1]/text()").get(),
                    "Bet %": match.xpath("./td[8]/div/span[1]/span/span/text()").get('').replace("%", ''),
                    "Money %": match.xpath("./td[9]/div/span[1]/div/text()").get('').replace("%", ''),
                })
                items.append({
                    "Date": match_date,
                    "Time": match_time,
                    "Game": match.xpath(
                        "./td[1]//div[@class='game-info__teams'][1]//div[@class='game-info__team--desktop']/span/text()").get() + ' vs ' + match.xpath(
                        "./td[1]//div[@class='game-info__teams'][2]//div[@class='game-info__team--desktop']/span/text()").get(),
                    "Team": match.xpath(
                        "./td[1]//div[@class='game-info__teams'][2]//div[@class='game-info__team--desktop']/span/text()").get(),
                    "Pro-Line": match.xpath("./td[3]/div/div[2]/text()").get(),
                    "Bet %": match.xpath("./td[8]/div/span[2]/span/span/text()").get('').replace("%", ''),
                    "Money %": match.xpath("./td[9]/div/span[2]/div/text()").get('').replace("%", ''),
                })
        except Exception as e:
            logger.info("error in parse_spread")
            logger.error(e)
        finally:
            return items

    def parse_total(self, response):
        items = []
        try:
            sel = Selector(text=response)
            for match in sel.xpath("//table/tbody/tr"):
                items.append({
                    "Team": match.xpath(
                        "./td[1]//div[@class='game-info__teams'][1]//div[@class='game-info__team--desktop']/span/text()").get(),
                    "O/U Pro-Line": match.xpath("./td[3]/div/div[1]/text()").get(),
                    "O/U Bet %": match.xpath("./td[8]/div/span[1]/span/span/text()").get('').replace("%", ''),
                    "O/U Money %": match.xpath("./td[9]/div/span[1]/div/text()").get('').replace("%", ''),
                })
                items.append({
                    "Team": match.xpath(
                        "./td[1]//div[@class='game-info__teams'][2]//div[@class='game-info__team--desktop']/span/text()").get(),
                    "O/U Pro-Line": match.xpath("./td[3]/div/div[2]/text()").get(),
                    "O/U Bet %": match.xpath("./td[8]/div/span[2]/span/span/text()").get('').replace("%", ''),
                    "O/U Money %": match.xpath("./td[9]/div/span[2]/div/text()").get('').replace("%", ''),
                })
        except Exception as e:
            logger.info("error in parse_total")
            logger.error(e)
        finally:
            return items

    def get_total(self, callback):
        try:
            self.page.locator("//div[@class='odds-tools-sub-nav__desktop-filter']").click()
            self.find_date()
            self.page.locator("//div[@data-testid='odds-tools-sub-nav__odds-type']/select").select_option(value="total")
            self.page.wait_for_timeout(5000)
            self.scroll_into_view("//div[@class='projections__table-container']")
            content = self.page.content()
            return callback(content)
        except Exception as e:
            logger.info("error in get_total")
            logger.error(e)

    def get_spread(self, callback):
        try:
            self.page.locator("//div[@class='odds-tools-sub-nav__desktop-filter']").click()
            self.find_date()
            self.scroll_into_view("//div[@class='projections__table-container']")
            content = self.page.content()
            return callback(content)
        except Exception as e:
            logger.info("error in get_spread")
            logger.error(e)

    def crawl(self):
        self.play, self.browser, self.page = init_playwright(False)
        combine_item = []
        try:
            if self.login():
                logger.info("Getting ActionNetwork")
                self.user_date = datetime.datetime.strptime(_date, "%Y-%m-%d")
                self.page.goto(self.url, timeout=playwright_timeout)
                modal_close = Selector(self.page.content()).xpath("//div[@class='modal__close']")
                if modal_close:
                    self.page.click("//div[@class='modal__close']")
                spread_items = self.get_spread(callback=self.parse_spread)
                total_items = self.get_total(callback=self.parse_total)
                # merge total_items to spread_items & create new var called combien_item
                for i in range(len(spread_items)):
                    for j in range(len(total_items)):
                        if spread_items[i]['Team'] == total_items[j]['Team']:
                            combine_item.append(spread_items[i] | total_items[j])
                            break
        except (KeyboardInterrupt, Exception):
            pass
        finally:
            self.browser.close()
            self.play.stop()
            return combine_item


crawler = ActionNetwork()
anet_combine = crawler.crawl()

###########################################


class Dime():

    def login(self):
        try:
            self.page.goto("https://www.sportsbettingdime.com/plus/register/", timeout=playwright_timeout)
            self.page.wait_for_timeout(5000)
            response = self.page.content()
            if is_exist(response, "//a[text()='PLUS']"):
                raise TimeoutError("Already logged in")
            self.page.locator("(//div[@class='d-flex flex-column align-items-center']/div[1]//input)[1]").fill("diego.lomanto@gmail.com")
            self.page.locator("(//div[@class='d-flex flex-column align-items-center']/div[1]//input)[2]").fill("v5gFU51^*0PO")
            self.page.wait_for_timeout(1000)
            self.page.locator("//div[@class='d-flex flex-column align-items-center']/div[1]//button").click()
            self.page.wait_for_selector("//a[text()='PLUS']", state="visible")
            logger.debug("logged in")
            return True
        except TimeoutError:
            return True
        except Exception as e:
            logger.debug("error in login")
            logger.error(e)
            return False
    
    
    def parse_total(self, response):
        items = []
        try:
            sel = Selector(text=response)
            for match in sel.xpath("//tr[@class='MuiTableRow-root']"):
                date_str = match.xpath("./th/div[contains(@class, 'rowShowDate')]/text()").get()
                match_date = dateparser.parse(date_str).date().strftime("%Y-%m-%d") if date_str else None
                if match_date == _date:
                    for team, name in zip(match.xpath(".//div[contains(@data-id, '-best') and not(@role)]"), match.xpath(".//div[contains(@class, 'nameTeam')]/text()").getall()):
                        total = team.xpath("./following-sibling::div[contains(@class, 'oddsText')]/div[contains(@class, 'first')]/text()").re_first("[0-9.]+")
                        over_under = team.xpath("./following-sibling::div[contains(@class, 'oddsText')]/div[contains(@class, 'first')]/text()").re_first("[a-z]")
                        best_odds = team.xpath("./following-sibling::div[contains(@class, 'oddsText')]/div[contains(@class, 'second')]/text()").re_first("[0-9+-]+")
                        book_icon = team.xpath("./following-sibling::div[contains(@class, 'bookLogoOuter')]/img/@src").get('').split('/')[-1]
                        items.append({
                            "Team": name,
                            "Over/Under": over_under,
                            "Total": total,
                            "Best Odds": best_odds.replace('-','') if best_odds and best_odds.strip() == '-' else best_odds,
                            "Book Icon": book_icon
                        })
        except Exception as e:
            logger.exception("Error parsing parse_total")
            logger.error(e)
        finally:
            return items
    

    def parse_spread(self, response):
        items = []
        try:
            sel = Selector(text=response)
            for match in sel.xpath("//tr[@class='MuiTableRow-root']"):
                date_str = match.xpath("./th/div[contains(@class, 'rowShowDate')]/text()").get()
                match_date = dateparser.parse(date_str).date().strftime("%Y-%m-%d") if date_str else None
                if match_date == _date:
                    for team, name in zip(match.xpath(".//div[contains(@data-id, '-best') and not(@role)]"), match.xpath(".//div[contains(@class, 'nameTeam')]/text()").getall()):
                        best_odds_spread = team.xpath("./following-sibling::div[contains(@class, 'oddsText')]/div[contains(@class, 'first')]/text()").re_first("[0-9.-]+")
                        best_odds_odds = team.xpath("./following-sibling::div[contains(@class, 'oddsText')]/div[contains(@class, 'second')]/text()").re_first("[0-9+-]+")
                        book_icon = team.xpath("./following-sibling::div[contains(@class, 'bookLogoOuter')]/img/@src").get('').split('/')[-1]
                        items.append({
                            "Team": name,
                            "Best Odds Spread": best_odds_spread.replace('-','') if best_odds_spread and best_odds_odds.strip() == '-' else best_odds_spread,
                            "Best Odds Odds": best_odds_odds.replace('-','') if best_odds_odds and best_odds_odds.strip() == '-' else best_odds_odds,
                            "Book Icon": book_icon
                        })
        except Exception as e:
            logger.exception("Error parsing parse_spread")
            logger.error(e)
        finally:
            return items


    def get_total(self, callback):
        try:
            self.page.get_by_label("spread").click()
            self.page.get_by_role("option", name="total").click()
            self.page.wait_for_timeout(15000)
            return callback(self.page.content())
        except Exception as e:
            logger.exception("Error getting get_total")
            logger.error(e)


    def get_spread(self, callback):
        try:
            self.page.wait_for_timeout(15000)
            return callback(self.page.content())
        except Exception as e:
            logger.exception("Error getting get_spread")
            logger.error(e)


    def crawl(self):
        self.play, self.browser, self.page = init_playwright()
        self.session = requests.Session()
        spread_items, total_items = [], []
        try:
            if self.login():
                logger.info("Getting Dime")
                cookies = playwright_cookies_to_requests(self.page.context.cookies())
                self.session.cookies = cookies
                self.page.goto("https://www.sportsbettingdime.com/nba/odds/", timeout=playwright_timeout)
                spread_items = self.get_spread(callback=self.parse_spread)
                total_items = self.get_total(callback=self.parse_total)
        except Exception:
            pass
        finally:
            self.browser.close()
            self.play.stop()
            return spread_items, total_items


# crawler = Dime()
# dime_spread, dime_total = crawler.crawl()


#############################################


class DE:
    start_url = "http://dokterentropy.com/users/dlomanto/preds1.NBA"


    def parse(self, response):
        items = []
        lines = response.text.splitlines()
        winner_idx = [i for i, line in enumerate(lines) if line.startswith("winner")][0]
        total_idx = [i for i, line in enumerate(lines) if line.startswith("totals")][0]
        slice = lines[winner_idx:total_idx]
        for line in slice:
            line = line.strip()
            if not line or 'winner' in line or '-----' in line:
                continue
            else:
                line = re.sub(r"\s\s+", "@@@", line)
                items.append({
                    "Winner": line.split("@@@")[0],
                    "By": "-" + re.search(r'[.0-9]+', line.split("@@@")[1]).group() if re.search(r'[.0-9]+', line.split("@@@")[1]) else "",
                    "Loser": re.search(r'[\sa-zA-Z]+', line.split("@@@")[1]).group(),
                    "%": line.split("@@@")[2],
                    "ML": line.split("@@@")[3],
                })
                items.append({
                    "Winner": re.search(r'[\sa-zA-Z]+', line.split("@@@")[1]).group(),
                    "By": re.search(r'[.0-9]+', line.split("@@@")[1]).group(),
                    "Loser": line.split("@@@")[0],
                    "%": 1 - float(line.split("@@@")[2].strip()),
                    "ML": "",
                })
        return items


    def get_page(self, callback):
        try:
            response = requests.get(self.start_url)
            return callback(response)
        except Exception as e:
            logger.debug("Error in get_page")
            logger.error(e)
            return []


    def crawl(self):
        logger.info("Getting DE")
        de_items = self.get_page(callback=self.parse)
        return de_items


crawler = DE()
de = crawler.crawl()


##########################################


class Sangarin:
    start_url = "http://sagarin.com/sports/nbasend.htm"

    def parse(self, response):
        sel = Selector(response.text)
        items = []
        lines = sel.xpath("//a[@name='Predictions_with_Totals']/parent::font/parent::h2/following-sibling::text()").get().splitlines()
        winner_idx = [i for i, line in enumerate(lines) if line.strip().startswith("FAVORITE")][0]
        total_idx = [i for i, line in enumerate(lines) if line.strip().startswith("EIGENVECTOR")][0]
        slice = lines[winner_idx:total_idx]
        for line in slice:
            if not line or 'FAVORITE' in line or '======' in line:
                continue
            else:
                line = re.sub(r"\s\s+", "@@@", line).strip('@@@')
                items.append({
                    "Favorite": line.split("@@@")[0],
                    "Rating": '-' + line.split("@@@")[1],
                    "Predict": '-' + line.split("@@@")[2],
                    "Golden": '-' + line.split("@@@")[3],
                    "Recent": '-' + line.split("@@@")[4],
                    "Underdog": line.split("@@@")[5],
                    "Odds": line.split("@@@")[6],
                    "PCT%": line.split("@@@")[7],
                    "Total": line.split("@@@")[8],
                })
                items.append({
                    "Favorite": line.split("@@@")[5],
                    "Rating": line.split("@@@")[1],
                    "Predict": line.split("@@@")[2],
                    "Golden": line.split("@@@")[3],
                    "Recent": line.split("@@@")[4],
                    "Underdog": '',
                    "Odds": '',
                    "PCT%": str(100 - int(line.split("@@@")[7].replace('%', ''))) + "%",
                    "Total": line.split("@@@")[8],
                })
        return items

    def get_page(self, url, callback):
        try:
            response = requests.get(url)
            return callback(response)
        except Exception as e:
            logger.debug("Error in get_page")
            logger.error(e)
            return []


    def crawl(self):
        logger.info("Getting Sangarin")
        sangarin_items = self.get_page(url=self.start_url, callback=self.parse)
        return sangarin_items


crawler = Sangarin()
sangarin = crawler.crawl()


#######################################


class Massey:


    def parse(self, response):
        sel = Selector(response)
        items = []
        for match in sel.xpath("//table[@class='mytable']/tbody/tr"):
            match_date = match.xpath("./td[@class='fdate sorted']/a/text()").get()
            match_time = match.xpath("./td[@class='fdate sorted']/div/text()").get()
            items.append({
                "Date": match_date,
                "Time": match_time,
                "Team": match.xpath("./td[2]/a/text()").get('').strip(),
                "Pred": match.xpath("./td[5]/text()").get(),
                "Pwin": match.xpath("./td[6]/text()").get(),
                "Spread": -(float(match.xpath("./td[5]/text()").get(0)) - float(match.xpath("./td[5]/div/text()").get(0))), # -(c2-c3)
                "Total": match.xpath("./td[8]/text()").get(),
            })
            items.append({
                "Date": match_date,
                "Time": match_time,
                "Team": match.xpath("./td[2]/div/a/text()").get('').replace("@",'').strip(),
                "Pred": match.xpath("./td[5]/div/text()").get(),
                "Pwin": match.xpath("./td[6]/div/text()").get(),
                "Spread": -(float(match.xpath("./td[5]/div/text()").get(0)) - float(match.xpath("./td[5]/text()").get(0))), # -(c3-c2)
                "Total": match.xpath("./td[8]/text()").get(),
            })
        return items


    def get_page(self, url, callback):
        try:
            self.page.goto(url, timeout=playwright_timeout)
            self.page.wait_for_selector("//table[@class='mytable']")
            response = self.page.content()
            return callback(response)
        except Exception as e:
            logger.debug("Error in get_page")
            logger.error(e)
            return []


    def crawl(self):
        self.play, self.browser, self.page = init_playwright()
        url = f"https://masseyratings.com/nba/games?dt={_date.replace('-','')}"
        logger.info("Getting Massey")
        massey_items = self.get_page(url=url, callback=self.parse)
        self.browser.close()
        self.play.stop()
        return massey_items


crawler = Massey()
massey = crawler.crawl()


###########################################


class Dratings:
    start_url = "https://www.dratings.com/predictor/nba-basketball-predictions/upcoming/{}"


    @staticmethod
    def cal_spread(match, i):
        try:
            if i == 1:
                return -round((float(match.xpath(f"./td[6]//text()").getall()[0]) - float(match.xpath(f"./td[6]//text()").getall()[-1])), 2)
            elif i == 2:
                return -round((float(match.xpath(f"./td[6]//text()").getall()[-1]) - float(match.xpath(f"./td[6]//text()").getall()[0])), 2)
        except Exception as e:
            return None


    def parse(self, response):
        items = []
        sel = Selector(response.text)
        for match in sel.xpath("//div[@id='scroll-upcoming']/table/tbody/tr"):
            for i in range(1,3):
                items.append({
                    "Time": match.xpath(".//time[@class='time-long']/@datetime").get('').split("T")[0],
                    "Teams": match.xpath(f"./td[2]/span[{i}]/a/text()").get(),
                    "Win": match.xpath(f"./td[3]/span[{i}]/text()").get(),
                    "Best ML": match.xpath(f"./td[4]/div[@class='vegas-sportsbook']/text()[{i}]").get(),
                    "Best Spread": match.xpath(f"./td[5]/div[@class='vegas-sportsbook']/text()[{i}]").get('').replace("Â",'').replace('½','.5'),
                    "Points": match.xpath(f"./td[6]/text()[{i}]").get(),
                    "Spread": self.cal_spread(match, i),
                    "Total Points": match.xpath(f"./td[7]/text()").get(),
                    "Best O/U": match.xpath(f"./td[8]/div[@class='vegas-sportsbook']/text()[{i}]").re_first("[ou]+"),
                    " ": match.xpath(f"./td[8]/div[@class='vegas-sportsbook']/text()[{i}]").get('').split('-')[0].replace("o", '').replace("u", '').replace("Â",'').replace('½','.5'),
                    "  ": match.xpath(f"./td[8]/div[@class='vegas-sportsbook']/text()[{i}]").get('').split('-')[-1].replace("o", '').replace("u", ''),
                    "Bet Value": match.xpath(f"./td[9]/div[@class='vegas-sportsbook']/text()[{i}]").get(),
                })
        return items
    

    @staticmethod
    def date_difference_in_days(date_str1, date_str2):
        date1 = datetime.datetime.strptime(date_str1, "%Y-%m-%d")
        date2 = datetime.datetime.strptime(date_str2, "%Y-%m-%d")
        date_difference = date2 - date1
        difference_in_days = date_difference.days + 1
        return difference_in_days  


    def get_page(self, url, callback):
        try:
            response = requests.get(url)
            return callback(response)
        except Exception as e:
            logger.debug("Error in get_page")
            logger.error(e)
            return []


    def crawl(self):
        logger.info("Getting Dratings")
        start_date = datetime.datetime.today().strftime("%Y-%m-%d")
        day = self.date_difference_in_days(start_date, _date)
        url = self.start_url.format(day)
        dratings_items = self.get_page(url, callback=self.parse)
        return dratings_items


crawler = Dratings()
dratings = crawler.crawl()


################################################


class BetQL():

    def login(self):
        try:
            self.page.goto("https://betql.co/nba/odds", timeout=playwright_timeout)
            self.page.wait_for_timeout(5000)
            response = self.page.content()
            if is_exist(response, "//div[@class='user-pic']"):
                raise TimeoutError("Already logged in")
            self.page.get_by_role("button", name="Log In").click()
            self.page.wait_for_timeout(1000)
            self.page.get_by_placeholder("Enter e-mail address").fill("diego.lomanto@gmail.com")
            self.page.get_by_placeholder("Enter password").fill(r"7!W?\\05o`v92")
            self.page.wait_for_timeout(1000)
            self.page.locator("//form[@class='rotoql-login__form']//button[text()='Log In']").click()
            self.page.wait_for_selector("//form[@class='rotoql-login__form']//button[text()='Log In']", state="hidden")
            logger.debug("logged in")
            return True
        except TimeoutError:
            return True
        except Exception as e:
            logger.debug("error in login")
            logger.error(e)
            return False
    
    
    def parse_match(self, response):
        items = []
        try:
            self.page.locator("//div[@id='total']").click()
            total_sel = Selector(self.page.content())
            total_rating = total_sel.xpath("//div[@class='rating-trend__title']/text()").get()

            sel = Selector(response)
            team1 = sel.xpath("//div[@class='team-header right']//div[@class='team-full-name']/text()").get()
            team2 = sel.xpath("//div[@class='team-header left']//div[@class='team-full-name']/text()").get()
            star_team = sel.xpath("//div[@class='rating-trend__favorite']/text()").get()

            for i in range(1, 3):
                if i == 1:
                    team = team1
                    score = sel.xpath("//div[@class='best-bets__main']/div[3]/div/div[contains(@class, 'projections-list')]/div[1]/span[1]/span/text()").get()
                    spread = -(float(score) - float(sel.xpath("//div[@class='best-bets__main']/div[3]/div/div[contains(@class, 'projections-list')]/div[1]/span[3]/span/text()").get(0))) # -(team1_score - team2_score)
                elif i == 2:
                    team = team2
                    score = sel.xpath("//div[@class='best-bets__main']/div[3]/div/div[contains(@class, 'projections-list')]/div[1]/span[3]/span/text()").get()
                    spread = -(float(score) - float(sel.xpath("//div[@class='best-bets__main']/div[3]/div/div[contains(@class, 'projections-list')]/div[1]/span[1]/span/text()").get(0))) # -(team2_score - team1_score)
                item = {
                    "Team": team,
                    "Score": score,
                    "Spread": spread,
                    "Rating": (total_rating.split('Rating')[0].strip() if total_rating else total_rating) if team.lower() in star_team.lower() else None,
                }
                items.append(item)
        except Exception as e:
            logger.debug("error in parse_odds")
            logger.error(e)
        finally:
            return items
    

    def get_match(self, url, callback):
        try:
            self.page.goto(url, timeout=playwright_timeout)
            self.page.mouse.wheel(1 ,1000)
            self.page.wait_for_selector("//div[@class='best-bets__main']/div[3]/div/div[contains(@class, 'projections-list')]")
            # self.page.wait_for_timeout(10000)
            response = self.page.content()
            return callback(response)
        except Exception as e:
            logger.debug("error in get_page")
            logger.error(e)


    def check_for_popup(self, popup):
        try:
            self.page.wait_for_selector(popup)
            return True
        except Exception as e:
            return False


    def get_spread(self):
        items = []
        try:
            logger.info("Getting BetQL")
            day = datetime.datetime.strptime(_date, "%Y-%m-%d").day
            self.page.wait_for_selector("//div[@class='games-container']")
            exist = self.check_for_popup("//button[contains(@class, 'modal-cancel-button')]") # 
            if exist:
                self.page.click("//button[contains(@class, 'modal-cancel-button')]")
            exist = self.check_for_popup("//button[@class='close']")
            if exist:
                self.page.click("//button[@class='close']")

            self.page.locator("//div[@class='d-none d-sm-flex games-view__filter-container']//button[contains(@class, 'rotoql-date-picker__button')]").click()
            self.page.wait_for_timeout(1000)
            self.page.locator(f"//div[@class='rotoql-date-picker__menu dropdown-menu show']//div[contains(@class,'rotoql-date-picker__calendar-cell ') and not(contains(@class, 'disabled'))]/span[text()='{day}']").click()
            self.page.wait_for_timeout(1000)
            self.page.wait_for_selector("//a[@class='games-table-column__team-link'][last()]")
            sel = Selector(self.page.content())
            for match in sel.xpath("//div[@class='games-table-column']/div/a"):
                url = urljoin("https://betql.co", match.xpath("./@href").get())
                item = self.get_match(url, callback=self.parse_match)
                items.extend(item)
        except Exception as e:
            logger.debug("error in get_odds")
            logger.error(e)
        finally:
            return items


    def crawl(self):
        self.play, self.browser, self.page = init_playwright()
        spread_items = []
        try:
            if self.login():
                spread_items = self.get_spread()
        except Exception:
            pass
        finally:
            self.browser.close()
            self.play.stop()
            return spread_items



crawler = BetQL()
betql_spread = crawler.crawl()


##########################################



class SportsLine():

    @staticmethod
    def minus_day_date(date_str):
        that_day = datetime.datetime.strptime(date_str, "%Y-%m-%d")
        yesterday = that_day - datetime.timedelta(days=1)
        return yesterday.strftime('%m/%d')


    @staticmethod
    def same_day_date(date_str):
        this_day = datetime.datetime.strptime(date_str, "%Y-%m-%d")
        return this_day.strftime('%m/%d')


    @staticmethod
    def uniform_format(date_str):
        date_str = date_str.split(",")[0]
        return datetime.datetime.strptime(date_str, '%b %d %Y').strftime('%m/%d')


    @staticmethod
    def extract_percentage(s):
        match = re.search(r'(\d+(\.\d+)?)%', s)
        if match:
            first = float(match.group(1))
            second = 100 - first
            return f"{first}%", f"{second}%"


    def login(self):
        try:
            self.page.goto("https://www.sportsline.com/login/", timeout=playwright_timeout)
            if is_exist(self.page.content(), "//a[contains(@da-tracking-nav, 'user-name-my-account')]"):
                raise TimeoutError("Already logged in")
            self.page.locator("//input[@id='loginId']").fill("diego.lomanto@gmail.com")
            self.page.locator("//input[@id='password']").fill("6rw6SgQMM#HZ4t$")
            self.page.locator("//button[text()='Log In']").click()
            self.page.wait_for_selector("//button[text()='Log In']", state="hidden")
            logger.debug("logged in")
            return True
        except TimeoutError:
            return True
        except Exception as e:
            logger.debug("error in login")
            logger.error(e)
            return False


    def get_pick_sheet(self, callback):
        try:
            self.page.goto("https://www.sportsline.com/nba/picks/", timeout=playwright_timeout)
            self.page.wait_for_selector("//div/main")
            response = self.page.content()
            return callback(response)
        except Exception as e:
            logger.debug("error in get_pick_sheet")
            logger.error(e)


    def parse_pick_sheet(self, response):
        items = []
        try:
            sel = Selector(response)
            for match in sel.xpath("//table/tbody/tr[not(@class)]"):
                teams = match.xpath("./td[1]//div[@data-testid='Team-name']/text()").getall()
                scores = [float(score) for score in match.xpath("./td[2]/div/div/text()").getall()]
                spreads = [-(scores[0] - scores[-1]), -(scores[-1] - scores[0])]
                totals = [sum(scores)] * 2
                for team, score, spread, total in zip(teams, scores, spreads, totals):
                    items.append({"Team": team, "Projected Score": score, "Spread": spread, "Total": total})
        except Exception as e:
            logger.debug("error in parse_pick_sheet")
            logger.error(e)
        finally:
            return items


    def parse_expert_sheet(self, response):
        items = []
        try:
            sel = Selector(response)
            that_day = self.minus_day_date(_date)
            this_day = self.same_day_date(_date)
            for expert in sel.xpath("//h1[contains(text(), 'Expert') and not(contains(text(), 'NBA'))]/following-sibling::section[contains(@data-id, 'STANDARD')]"): # latest
            # for expert in sel.xpath("//h1[contains(text(), 'Past') and not(contains(text(), 'NBA'))]/following-sibling::section[contains(@data-id, 'STANDARD')]"): # past picks
                if f"{that_day}" in self.uniform_format(expert.xpath(".//span[contains(text(), '202')]/text()").get('')): # from top to bottom, break when hit yesterday
                    break
                team = expert.xpath(".//span[contains(text(), 'Point') or contains(text(), 'Over')]/following-sibling::span/text()").re_first(r"[a-zA-z\s\.]+")
                if 'over' in team.lower() or 'under' in team.lower():
                    if 'over' in team.lower():
                        team = expert.xpath(".//span[contains(text(), '@')]/parent::div/parent::div/preceding-sibling::div/div//span[@color]/text()").get()
                    elif 'under' in team.lower():
                        team = expert.xpath(".//span[contains(text(), '@')]/text()").get('').replace("@",'')
                pick = expert.xpath(".//span[contains(text(), 'Point') or contains(text(), 'Over')]/following-sibling::span/text()").re_first(r"[-\+\d]+")
                analysis = expert.xpath(".//span[contains(text(), 'Analysis:')]/following-sibling::p/text()").get()
                expert_name = expert.xpath(".//a[@data-tracking-value='expert-picks_click_profile-avatar']/div/div[@direction]/span[1]/text()").get('')
                if f"{this_day}" in self.uniform_format(expert.xpath(".//span[contains(text(), '202')]/text()").get('')):
                    items.append({
                        "Team": team,
                        "Pick": pick,
                        "Analysis": analysis,
                        "Expert": expert_name
                    })
        except Exception as e:
            logger.debug("error in parse_expert_sheet")
            logger.error(e)
        finally:
            return items


    def get_expert_sheet(self, url, callback):
        try:
            self.page.goto(url, timeout=playwright_timeout)
            self.page.wait_for_selector("//div/main")
            for _ in range(10):
                response = self.page.content()
                sel = Selector(response)
                if sel.xpath("//div[@data-testid='loadmore-container']"):
                    self.page.locator("//div[@data-testid='loadmore-container']").scroll_into_view_if_needed()
                    self.page.locator("//div[@data-testid='loadmore-container']/button").click()
                    self.page.wait_for_timeout(1000)
                else:
                    break
            return callback(response)
        except Exception as e:
            logger.debug("error in get_expert_sheet_spread")
            logger.error(e)


    def crawl(self):
        self.play, self.browser, self.page = init_playwright()
        self.page.context.route("**/*", lambda route: route.abort() if route.request.resource_type in ["image", "media", "font"] else route.continue_())
        pick_sheet, expert_sheet = [], []
        try:
            if self.login():
                logger.info("Getting SportsLine")
                pick_sheet = self.get_pick_sheet(callback=self.parse_pick_sheet)
                expert_sheet_spread_items = self.get_expert_sheet("https://www.sportsline.com/nba/picks/experts/?pickType=POINT_SPREAD", callback=self.parse_expert_sheet)
                expert_sheet_ou_items = self.get_expert_sheet("https://www.sportsline.com/nba/picks/experts/?pickType=OVER_UNDER", callback=self.parse_expert_sheet)
                expert_sheet = expert_sheet_spread_items + expert_sheet_ou_items
        except Exception:
            pass
        finally:
            self.browser.close()
            self.play.stop()
            return pick_sheet, expert_sheet


crawler = SportsLine()
pick_sheet, expert_sheet = crawler.crawl()

#######################################



class PT:
    start_url = "https://www.thepredictiontracker.com/prednba.php"


    @staticmethod
    def opposite_sign(item):
        for key, value in item.items():
            if value is not None and value and key not in ["Home", "Road"]:
                if value.startswith("-"):
                    item[key] = value.replace("-", "+")
                elif value.startswith("+") or not value.startswith('-'):
                    item[key] = "-" + value.replace("+", "-")
        item['Home'], item['Road'] = item['Road'], item['Home']
        return item


    def parse(self, response):
        items = []
        sel = Selector(response.text)
        for row in sel.xpath("//table[last()]/tbody/tr"):
            item = {
                "Home": row.xpath("td[1]/text()").get(),
                "Road": row.xpath("td[2]/text()").get(),
                "Line": row.xpath("td[3]/text()").get(),
                "lineopen": row.xpath("td[4]/text()").get(),
                "Sagarin": row.xpath("td[5]/text()").get(),
                "Sagarin Recent": row.xpath("td[6]/text()").get(),
                "Sagarin Pred.": row.xpath("td[7]/text()").get(),
                "Sagarin Golden Mean": row.xpath("td[8]/text()").get(),
                "Sonny Moore": row.xpath("td[9]/text()").get(),
                "Massey Ratings": row.xpath("td[10]/text()").get(),
                "Stat Fox": row.xpath("td[11]/text()").get(),
                "ESPN FPI": row.xpath("td[12]/text()").get(),
                "Dunkel Index": row.xpath("td[13]/text()").get(),
                "Dokter Entropy": row.xpath("td[14]/text()").get(),
                "Versus Sports Simulator": row.xpath("td[15]/text()").get(),
                "Donchess Inference": row.xpath("td[16]/text()").get(),
                "Talisman Red": row.xpath("td[17]/text()").get(),
                "Roundtable": row.xpath("td[18]/text()").get(),
                "lineavg": row.xpath("td[19]/text()").get(),
            }
            # change postive to negative and vice versa
            items.append(self.opposite_sign(copy(item)))
            items.append(item)
        return items


    def get_page(self, callback):
        try:
            response = requests.get(self.start_url)
            return callback(response)
        except Exception as e:
            logger.error(f"Error in get_page: {e}")
            return []


    def crawl(self):
        logger.info("Getting PT")
        pt_items = self.get_page(callback=self.parse)
        return pt_items


crawler = PT()
pt = crawler.crawl()


##########################################


class BettingPros:


    def parse(self, response):
        teams = []
        if response:
            sel = Selector(response)
            for match in sel.xpath("//div[@class='odds-offers__tables-container']/section/div[@style and position()>1]/div"):
                name1 = match.xpath(".//div[@class='odds-offer-label']/div[1]//a/@href").get('').strip('/').split('/')[-1].capitalize()
                name2 = match.xpath(".//div[@class='odds-offer-label']/div[2]//a/@href").get('').strip('/').split('/')[-1].capitalize()
                team1 = {
                    "Team": name1,
                    "Open": match.xpath(".//div[@class='odds-offer__item odds-offer__item--open']/div[1]//span[1]/text()").get(),
                    "Spread": match.xpath(".//div[@class='odds-offer__item odds-offer__item--best-odds']/button[1]//span[1]/text()").get(),
                    "Line": match.xpath(".//div[@class='odds-offer__item odds-offer__item--best-odds']/button[1]//span[2]/text()").get('').replace('(','').replace(')','')
                }
                team2 = {
                    "Team": name2,
                    "Open": match.xpath(".//div[@class='odds-offer__item odds-offer__item--open']/div[2]//span[1]/text()").get(),
                    "Spread": match.xpath(".//div[@class='odds-offer__item odds-offer__item--best-odds']/button[2]//span[1]/text()").get(),
                    "Line": match.xpath(".//div[@class='odds-offer__item odds-offer__item--best-odds']/button[2]//span[2]/text()").get('').replace('(','').replace(')','')
                }
                teams.append(team1)
                teams.append(team2)
        return teams


    def parse_total(self, response):
        teams = []
        if response:
            sel = Selector(response)
            for match in sel.xpath("//div[@class='odds-offers__tables-container']/section/div[@style and position()>1]/div"):
                name1 = match.xpath(".//div[@class='odds-offer-label']/div[1]//a/@href").get('').strip('/').split('/')[-1].capitalize()
                name2 = match.xpath(".//div[@class='odds-offer-label']/div[2]//a/@href").get('').strip('/').split('/')[-1].capitalize()
                team1 = {
                    "Team": name1,
                    "Open": match.xpath(".//div[@class='odds-offer__item odds-offer__item--open']/div[1]//span[1]/text()").re_first('[0-9.+]+'),
                    "O/U": match.xpath(".//div[@class='odds-offer__item odds-offer__item--best-odds']/button[1]//span[1]/text()").re_first(r'\w+'),
                    "Total": match.xpath(".//div[@class='odds-offer__item odds-offer__item--best-odds']/button[1]//span[1]/text()").re_first('[0-9.+]+'),
                    "Line": match.xpath(".//div[@class='odds-offer__item odds-offer__item--best-odds']/button[1]//span[2]/text()").get('').replace('(','').replace(')','')
                }
                team2 = {
                    "Team": name2,
                    "Open": match.xpath(".//div[@class='odds-offer__item odds-offer__item--open']/div[2]//span[1]/text()").re_first('[0-9.+]+'),
                    "O/U": match.xpath(".//div[@class='odds-offer__item odds-offer__item--best-odds']/button[2]//span[1]/text()").re_first(r'\w+'),
                    "Total": match.xpath(".//div[@class='odds-offer__item odds-offer__item--best-odds']/button[2]//span[1]/text()").re_first('[0-9.+]+'),
                    "Line": match.xpath(".//div[@class='odds-offer__item odds-offer__item--best-odds']/button[2]//span[2]/text()").get('').replace('(','').replace(')','')
                }
                teams.append(team1)
                teams.append(team2)
        return teams



    def exist(self, selector):
        try:
            self.page.wait_for_selector(selector)
            return True
        except TimeoutError as e:
            return False


    def get_page(self, url, callback, _type):
        try:
            self.page.goto(url, timeout=playwright_timeout)
            if not self.exist("//div[@class='odds-offers__tables-container']"):
                response = None
            else:
                if _type == "odds":
                    self.page.click("//a[text()='Totals']")
                    self.page.wait_for_timeout(10000)
                    
                    self.page.click("//a[text()='Point Spread']")
                    self.page.wait_for_timeout(10000)
                elif _type == "total":
                    self.page.click("//a[text()='Point Spread']")
                    self.page.wait_for_timeout(10000)
                    
                    self.page.click("//a[text()='Totals']")
                    self.page.wait_for_timeout(10000)

                self.page.mouse.wheel(0,1000)
                self.page.wait_for_timeout(60000)
                response = self.page.content()
            return callback(response)
        except Exception as e:
            logger.error(f"Error in get_page: {e}")
            return []


    def crawl(self):
        self.play, self.browser ,self.page = init_playwright()
        logger.info("Getting BettingPros")
        bettingpros_odds_items = self.get_page(f"https://www.bettingpros.com/nba/odds/?date={_date}", callback=self.parse, _type="odds")
        bettingpros_total_items = self.get_page(f"https://www.bettingpros.com/nba/odds/total/?date={_date}", callback=self.parse_total, _type="total")
        self.browser.close()
        self.play.stop()
        return bettingpros_odds_items, bettingpros_total_items


crawler = BettingPros()
bettingpros_odds_items, bettingpros_total_items = crawler.crawl()



#########################################



class PayDirt():
    paydirt_url = "https://paydirtdfs.com/nba-paywalled/nba-prop-betting-tool/"
    am_url = "https://paydirtdfs.com/nba-paywalled/americannumbers-nba-team-level-models/"


    def wait_for(self, frame, selector, timeout=10000):
        try:
            frame.wait_for_selector(selector, timeout=timeout)
            return True
        except Exception as e:
            # logger.error(e)
            return False


    def login(self):
        try:
            self.page.goto("https://paydirtdfs.com/login/", timeout=playwright_timeout)
            self.wait_for(self.page, "//input[@id='user_login']", timeout=playwright_timeout)
            response = self.page.content()
            if not is_exist(response, "//input[@id='user_login']"):
                raise TimeoutError("Already logged in")
            self.page.locator("//input[@id='user_login']").fill("TPERFB")
            self.page.wait_for_timeout(1000)
            self.page.locator("//input[@id='user_pass']").fill("testing_password_111")
            self.page.wait_for_timeout(1000)
            self.page.locator("//input[@id='rememberme']").check()
            self.page.wait_for_timeout(1000)
            self.page.locator("//input[@id='wp-submit']").click()
            self.page.wait_for_selector("//input[@id='wp-submit']", state="hidden")
            logger.debug("logged in")
            return True
        except TimeoutError as e:
            logger.debug(e)
            return True
        except Exception as e:
            logger.debug("error in login")
            logger.error(e)
            return False
        

    def handle_route(self, route: Route):
        logger.info("Intercepted CSV response")
        response = route.fetch()
        self.captured_df = pd.read_csv(io.StringIO(response.text()))


    def get_am(self):
        try:
            self.page.goto(self.am_url, timeout=playwright_timeout)
            self.wait_for(self.page, "//iframe")
            iframe = self.page.frames[1]
            exists = self.wait_for(iframe, "//p[contains(text(), 'Export Game Model')]/ancestor::button", timeout=playwright_timeout)
            if exists:
                iframe.click("//p[contains(text(), 'Export Game Model')]/ancestor::button")
            iframe.wait_for_timeout(10000)
            self.captured_df['Date'] = pd.to_datetime(self.captured_df['Date'], format='%a, %b %d, %Y')
            target_date = datetime.datetime.strptime(_date, "%Y-%m-%d").date()
            filtered_df = self.captured_df[self.captured_df['Date'].dt.date == target_date]
            self.csv_data = filtered_df.to_dict()
            self.csv_data.pop("Unnamed: 0")
            return self.csv_data
        except Exception as e:
            logger.info("error in get_am")
            logger.error(e)


    def get_paydirt(self):
        try:
            self.page.goto(self.paydirt_url, timeout=playwright_timeout)
            self.wait_for(self.page, "//iframe")
            iframe = self.page.frames[1]
            exists = self.wait_for(iframe, "//p[contains(text(), 'Export Team Model')]/ancestor::button", timeout=playwright_timeout)
            if exists:
                iframe.click("//p[contains(text(), 'Export Team Model')]/ancestor::button")
                iframe.wait_for_timeout(2000)
                self.csv_data = self.captured_df.to_dict()
                return self.csv_data
        except Exception as e:
            logger.info("error in get_paydirt")
            logger.error(e)


    def crawl(self):
        self.play, self.browser, self.page = init_playwright()
        am_items, paydirt_items = None, None
        self.page.route("**/*.csv", self.handle_route)
        try:
            if self.login():
                logger.info("Getting PayDirt")
                am_items = self.get_am()
                paydirt_items = self.get_paydirt()
        except (KeyboardInterrupt, Exception):
            pass
        finally:
            self.browser.close()
            self.play.stop()
            return (am_items, paydirt_items)

crawler = PayDirt()
am_items, paydirt_items = crawler.crawl()



#########################################



class Nba():
    team_url = "https://www.nba.com/stats/teams/advanced?LastNGames=10"
    home_url = "https://www.nba.com/stats/teams/advanced?LastNGames=20&Location=Home"
    road_url = "https://www.nba.com/stats/teams/advanced?LastNGames=20&Location=Road"


    def wait_for(self, selector, timeout=10000):
        try:
            self.page.wait_for_selector(selector, timeout=timeout)
            return True
        except Exception as e:
            logger.error(e)
            return False
        

    def parse(self, response):
        items = []
        sel = Selector(response)
        headers = [h.upper() for h in sel.xpath("//table[not(contains(@class, 'Date'))]/thead/tr/th[not(@hidden)]//text()").getall()]
        for row in sel.xpath("//table[not(contains(@class, 'Date'))]/tbody/tr"):
            item = {}
            for header, col in zip(headers, row.xpath("./td")):
                item[header] = "".join(col.xpath(".//text()").getall()).strip()
            items.append(item)
        return items
    

    def get_team(self, callback=None):
        try:
            self.page.goto(self.team_url, timeout=playwright_timeout)
            self.wait_for("//table[not(contains(@class, 'Date'))]/tbody", timeout=playwright_timeout)
            content = self.page.content()
            return callback(content)
        except Exception as e:
            logger.info("error in get_team")
            logger.error(e)


    def get_home(self, callback=None):
        try:
            self.page.goto(self.home_url, timeout=playwright_timeout)
            self.wait_for("//table[not(contains(@class, 'Date'))]/tbody", timeout=playwright_timeout)
            content = self.page.content()
            return callback(content)
        except Exception as e:
            logger.info("error in get_home")
            logger.error(e)

    
    def get_road(self, callback=None):
        try:
            self.page.goto(self.road_url, timeout=playwright_timeout)
            self.wait_for("//table[not(contains(@class, 'Date'))]/tbody", timeout=playwright_timeout)
            content = self.page.content()
            return callback(content)
        except Exception as e:
            logger.info("error in get_road")
            logger.error(e)


    def crawl(self):
        self.play, self.browser, self.page = init_playwright()
        team_items, home_items, road_items = None, None, None
        try:
            logger.info("Getting NBA")
            team_items = self.get_team(callback=self.parse)
            home_items = self.get_home(callback=self.parse)
            road_items = self.get_road(callback=self.parse)
        except (KeyboardInterrupt, Exception):
            pass
        finally:
            self.browser.close()
            self.play.stop()
            return (team_items, home_items, road_items)
        

crawler = Nba()
team_items, home_items, road_items = crawler.crawl()



#########################################################
    

def dump():
    logger.info("Saving")
    replace_Stings = [None]
    replace_with = "TBD"
    
    df1 = pd.DataFrame(paydirt_items).replace(replace_Stings, replace_with)
    df2 = pd.DataFrame(am_items).replace(replace_Stings, replace_with)
    df3 = pd.DataFrame(anet_combine).replace(replace_Stings, replace_with)
    # df2 = pd.DataFrame(dime_spread).replace(replace_Stings, replace_with)
    df4 = pd.DataFrame(bettingpros_odds_items).replace(replace_Stings, replace_with)
    # df3 = pd.DataFrame(dime_total).replace(replace_Stings, replace_with)
    df5 = pd.DataFrame(bettingpros_total_items).replace(replace_Stings, replace_with)
    df6 = pd.DataFrame(pt).replace(replace_Stings, replace_with)
    df7 = pd.DataFrame(de).replace(replace_Stings, replace_with)
    df8 = pd.DataFrame(sangarin).replace(replace_Stings, replace_with)
    df9 = pd.DataFrame(massey).replace(replace_Stings, replace_with)
    df10 = pd.DataFrame(dratings).replace(replace_Stings, replace_with)
    df11 = pd.DataFrame(betql_spread).replace(replace_Stings, replace_with)
    df12 = pd.DataFrame(pick_sheet).replace(replace_Stings, replace_with)
    df13 = pd.DataFrame(expert_sheet).replace(replace_Stings, replace_with)
    df14 = pd.DataFrame(team_items).replace(replace_Stings, replace_with)
    df15 = pd.DataFrame(home_items).replace(replace_Stings, replace_with)
    df16 = pd.DataFrame(road_items).replace(replace_Stings, replace_with)


    with pd.ExcelWriter(f'master_NBA_file.xlsx', engine='openpyxl') as writer:
        df1.to_excel(writer, sheet_name="Paydirt", index=False)
        df2.to_excel(writer, sheet_name="AM", index=False)
        df3.to_excel(writer, sheet_name="ActPubSPR", index=False)
        # df2.to_excel(writer, sheet_name="SP Spreads", index=False)
        df4.to_excel(writer, sheet_name="BPSpread", index=False)
        # df3.to_excel(writer, sheet_name="SPD Total", index=False)
        df5.to_excel(writer, sheet_name="BPTotal", index=False)
        df6.to_excel(writer, sheet_name="PT", index=False)
        df7.to_excel(writer, sheet_name="DE", index=False)
        df8.to_excel(writer, sheet_name="Sangarin", index=False)
        df9.to_excel(writer, sheet_name="Massey", index=False)
        df10.to_excel(writer, sheet_name="DRatings", index=False)
        df11.to_excel(writer, sheet_name="BetQL", index=False)
        df12.to_excel(writer, sheet_name="Sportsline", index=False)
        df13.to_excel(writer, sheet_name="SPExperts", index=False)
        df14.to_excel(writer, sheet_name="TeamStats", index=False)
        df15.to_excel(writer, sheet_name="HomeStats", index=False)
        df16.to_excel(writer, sheet_name="RoadStats", index=False)

dump()
convert_cells('master_NBA_file.xlsx')
logger.info(f"Creating Spreadsheet")
to_template(_date) 
